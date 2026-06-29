import io
import xml.etree.ElementTree as ET
import streamlit as st
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ============================================
# Configuração da página
# ============================================
st.set_page_config(
    page_title="CIOT - Validador de Coordenadas",
    page_icon="📦",
    layout="wide"
)

# Esconde elementos padrão do Streamlit
st.markdown("""
    <style>
        #MainMenu {visibility: hidden;}
        footer {visibility: hidden;}
        header {visibility: hidden;}
        [data-testid="stToolbar"] {visibility: hidden;}
        [data-testid="manage-app-button"] {display: none;}
        .viewerBadge_container__r5tak {display: none;}
        .viewerBadge_link__qRIco {display: none;}
    </style>
""", unsafe_allow_html=True)

st.title("📦 CIOT — Validador de Coordenadas")
st.markdown("Faça o upload do arquivo XML para validar as coordenadas dos CT-es.")

# ============================================
# Upload do arquivo
# ============================================
arquivo = st.file_uploader("Selecione o arquivo XML", type=["xml"])

if arquivo is None:
    st.info("Aguardando upload do arquivo XML...")
    st.stop()

# ============================================
# Namespace
# ============================================
ns = {
    "adic": "http://schemas.ipc.adm.br/efrete/pefV2/AdicionarOperacaoTransporte"
}

try:
    tree = ET.parse(arquivo)
    root = tree.getroot()
except ET.ParseError:
    st.error("❌ Erro ao ler o XML. Verifique se o arquivo está correto.")
    st.stop()

# ============================================
# Processa os CT-es
# ============================================
ctes_com_erro = []
linhas_tabela = []
todas_viagens = root.findall(".//adic:Viagens", ns)
total = len(todas_viagens)

with st.spinner("Processando CT-es..."):

    wb = Workbook()
    ws = wb.active
    ws.title = "Validação"

    cabecalho = [
        "CT-e",
        "Latitude Origem",
        "Longitude Origem",
        "Latitude Destino",
        "Longitude Destino",
        "Campos em branco"
    ]
    ws.append(cabecalho)

    azul = PatternFill(fill_type="solid", fgColor="1F4E78")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = azul
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.freeze_panes = "A2"

    for viagem in todas_viagens:

        cte = viagem.findtext(
            "adic:DocumentoViagem", default="", namespaces=ns
        ).replace("CTe-", "").strip()

        lat_ori  = viagem.findtext("adic:LatitudeOrigem",   default="", namespaces=ns)
        lon_ori  = viagem.findtext("adic:LongitudeOrigem",  default="", namespaces=ns)
        lat_dest = viagem.findtext("adic:LatitudeDestino",  default="", namespaces=ns)
        lon_dest = viagem.findtext("adic:LongitudeDestino", default="", namespaces=ns)

        erros = []

        origem_faltante  = not lat_ori.strip()  and not lon_ori.strip()
        lat_ori_faltante = not lat_ori.strip()  and lon_ori.strip()
        lon_ori_faltante = lat_ori.strip()      and not lon_ori.strip()

        destino_faltante  = not lat_dest.strip() and not lon_dest.strip()
        lat_dest_faltante = not lat_dest.strip() and lon_dest.strip()
        lon_dest_faltante = lat_dest.strip()     and not lon_dest.strip()

        if origem_faltante:
            erros.append("Coordenadas do Remetente Ausentes")
        else:
            if lat_ori_faltante:
                erros.append("Latitude do Remetente Ausente")
            if lon_ori_faltante:
                erros.append("Longitude do Remetente Ausente")

        if destino_faltante:
            erros.append("Coordenadas do Destinatário Ausentes")
        else:
            if lat_dest_faltante:
                erros.append("Latitude do Destinatário Ausente")
            if lon_dest_faltante:
                erros.append("Longitude do Destinatário Ausente")

        if erros:
            descricao_erro = ", ".join(erros)
            ws.append([cte, lat_ori, lon_ori, lat_dest, lon_dest, descricao_erro])
            ctes_com_erro.append(cte)
            linhas_tabela.append({
                "CT-e": cte,
                "Latitude Origem": lat_ori if lat_ori.strip() else "—",
                "Longitude Origem": lon_ori if lon_ori.strip() else "—",
                "Latitude Destino": lat_dest if lat_dest.strip() else "—",
                "Longitude Destino": lon_dest if lon_dest.strip() else "—",
                "Problema": descricao_erro
            })

    # Centraliza células
    for linha in ws.iter_rows():
        for celula in linha:
            celula.alignment = Alignment(horizontal="center", vertical="center")

    # Linha do KMM
    linha_kmm = ws.max_row + 4
    lista_ctes = ",".join(ctes_com_erro)

    titulo = ws.cell(row=linha_kmm, column=1)
    titulo.value = "Copiar e colar no KMM >>"
    titulo.font = Font(bold=True, color="FFFFFF")
    titulo.fill = azul
    titulo.alignment = Alignment(horizontal="center")

    celula_kmm = ws.cell(row=linha_kmm, column=2)
    celula_kmm.value = lista_ctes
    celula_kmm.font = Font(bold=True)
    celula_kmm.fill = PatternFill(fill_type="solid", fgColor="FFF200")
    celula_kmm.alignment = Alignment(horizontal="center")

    # Largura das colunas
    for coluna in ws.columns:
        maior = 0
        letra = get_column_letter(coluna[0].column)
        for c in coluna:
            if c.value is not None:
                maior = max(maior, len(str(c.value)))
        ws.column_dimensions[letra].width = maior + 3

    ws.column_dimensions["B"].width = max(
        ws.column_dimensions["B"].width,
        min(len(lista_ctes) * 0.9, 120)
    )

    ultima_linha_tabela = len(ctes_com_erro) + 1
    ws.auto_filter.ref = f"A1:F{ultima_linha_tabela}"

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

# ============================================
# Exibe resultado
# ============================================
st.divider()

if ctes_com_erro:
    col1, col2 = st.columns(2)
    col1.metric("CT-es analisados", total)
    col2.metric("CT-es com erro", len(ctes_com_erro), delta=f"-{len(ctes_com_erro)}", delta_color="inverse")

    st.error(f"⚠️ {len(ctes_com_erro)} CT-e(s) com coordenadas ausentes encontrados.")

    # Tabela visual na tela
    st.subheader("📋 CT-es com problema")
    df = pd.DataFrame(linhas_tabela)
    st.dataframe(
        df,
        use_container_width=True,
        hide_index=True,
        column_config={
            "CT-e": st.column_config.TextColumn("CT-e", width="small"),
            "Latitude Origem": st.column_config.TextColumn("Lat. Origem"),
            "Longitude Origem": st.column_config.TextColumn("Long. Origem"),
            "Latitude Destino": st.column_config.TextColumn("Lat. Destino"),
            "Longitude Destino": st.column_config.TextColumn("Long. Destino"),
            "Problema": st.column_config.TextColumn("Problema", width="large"),
        }
    )

    # KMM
    st.subheader("🔗 Lista para o KMM")
    st.code(lista_ctes, language=None)
    st.caption("Copie e cole diretamente no KMM.")

    # Download Excel
    st.divider()
    nome_saida = arquivo.name.replace(".xml", "_Validacao_Coordenadas.xlsx")
    st.download_button(
        label="⬇️ Baixar Excel com resultado",
        data=buffer,
        file_name=nome_saida,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

else:
    st.metric("CT-es analisados", total)
    st.success("✅ Nenhum CT-e com coordenadas em branco foi encontrado!")
